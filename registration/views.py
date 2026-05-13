from django.shortcuts import render
from django.views.generic import CreateView, ListView, DetailView, TemplateView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils.decorators import method_decorator
from django_ratelimit.decorators import ratelimit

from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action

from .models import Registration, RegistrationStats
from .serializers import RegistrationSerializer, StatsSerializer
from .forms import RegistrationForm

# --- UI Views ---

class IndexView(TemplateView):
    template_name = 'index.html'

@method_decorator(ratelimit(key='ip', rate='5/h', method='POST', block=True), name='post')
class RegistrationCreateView(CreateView):
    model = Registration
    form_class = RegistrationForm
    template_name = 'registration_form.html'
    success_url = reverse_lazy('registration_success')

    def form_valid(self, form):
        # Honeypot check — if the hidden field has a value, it's a bot
        if self.request.POST.get('website_url', ''):
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden('Spam detected.')
        return super().form_valid(form)

class RegistrationSuccessView(TemplateView):
    template_name = 'registration_success.html'

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'admin_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate stats dynamically
        context['stats'] = {
            'total_registrations': Registration.objects.count(),
            'approved': Registration.objects.filter(status='A').count(),
            'pending': Registration.objects.filter(status='P').count(),
            'rejected': Registration.objects.filter(status='R').count(),
            'drivers': Registration.objects.filter(category='D').count(),
            'navigators': Registration.objects.filter(category='N').count(),
            'male': Registration.objects.filter(gender='M').count(),
            'female': Registration.objects.filter(gender='F').count(),
        }
        
        context['recent_registrations'] = Registration.objects.all()[:10]
        return context

class RegistrationListView(LoginRequiredMixin, ListView):
    model = Registration
    template_name = 'registration_list.html'
    context_object_name = 'registrations'
    paginate_by = 20

class RegistrationDetailView(LoginRequiredMixin, DetailView):
    model = Registration
    template_name = 'registration_detail.html'
    context_object_name = 'registration'


class ExportExcelView(LoginRequiredMixin, TemplateView):
    """Export all registrations to Excel file"""

    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from django.utils import timezone

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Registrations'

        # Styles
        header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E8ECF1'),
            right=Side(style='thin', color='E8ECF1'),
            top=Side(style='thin', color='E8ECF1'),
            bottom=Side(style='thin', color='E8ECF1'),
        )

        # Column headers
        headers = [
            '#', 'الاسم الكامل', 'العمر', 'الجنس', 'رقم الهوية', 'الجنسية',
            'رقم الجوال', 'البريد الإلكتروني', 'الفئة', 'رخصة القيادة',
            'الرخصة الرياضية', 'رقم الرخصة الرياضية', 'خبرة سابقة',
            'أنواع السباقات', 'عدد السباقات', 'سنوات الخبرة',
            'تقييم القيادة الصحراوية', 'الهدف من البرنامج',
            'الالتزام بـ 5 أيام', 'الحالات الصحية', 'منطقة الإقامة',
            'القدرة على السفر', 'حالة الطلب', 'تاريخ التسجيل',
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        status_map = dict(Registration.STATUS_CHOICES)
        registrations = Registration.objects.all().order_by('-application_date')

        for row_idx, reg in enumerate(registrations, 2):
            data = [
                row_idx - 1,
                reg.full_name,
                reg.age,
                reg.get_gender_display(),
                reg.id_number,
                reg.nationality,
                reg.phone,
                reg.email,
                reg.get_category_display(),
                'نعم' if reg.has_driving_license else 'لا',
                'نعم' if reg.has_sports_license else 'لا',
                reg.sports_license_number or '-',
                'نعم' if reg.has_previous_experience else 'لا',
                ', '.join(reg.race_types) if reg.race_types else '-',
                reg.number_of_races or '-',
                reg.get_years_of_experience_display() if reg.years_of_experience else '-',
                reg.desert_driving_rating,
                reg.program_goal,
                'نعم' if reg.committed_to_5_days else 'لا',
                reg.health_conditions or '-',
                reg.get_residence_region_display(),
                'نعم' if reg.can_travel else 'لا',
                str(status_map.get(reg.status, reg.status)),
                reg.application_date.strftime('%Y-%m-%d %H:%M') if reg.application_date else '-',
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Build response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'registrations_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# --- API ViewSets ---

class RegistrationViewSet(viewsets.ModelViewSet):
    queryset = Registration.objects.all()
    serializer_class = RegistrationSerializer
    
    def get_permissions(self):
        if self.action == 'create':
            return [permissions.AllowAny()]
        return [permissions.IsAdminUser()]

    @action(detail=True, methods=['patch'])
    def status(self, request, pk=None):
        registration = self.get_object()
        new_status = request.data.get('status')
        if new_status in dict(Registration.STATUS_CHOICES):
            registration.status = new_status
            registration.save()
            return Response({'status': 'status updated'})
        return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

class StatsViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = RegistrationStats.objects.all()
    serializer_class = StatsSerializer
    permission_classes = [permissions.IsAdminUser]
